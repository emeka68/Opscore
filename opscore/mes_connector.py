"""
MES Connector Module
Import data FROM and export data TO Manufacturing Execution / WMS / ERP systems.

Supported formats:
  - CSV  (universal — SAP, Oracle, Infor, Epicor, NetSuite, QuickBooks)
  - JSON (REST-based MES/WMS APIs)
  - Excel XLSX (most desktop-based MES exports)
  - XML  (SAP iDocs, EDI 856 ASN, legacy WMS)

Import targets: shipments, kpis, inventory, orders
Export targets: opscore reports, anomaly summaries, SOP documents
"""

import csv
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

SUPPORTED_IMPORT = ["csv", "json", "xlsx", "xml"]
SUPPORTED_EXPORT = ["csv", "json", "xlsx", "xml"]

# ── IMPORT ─────────────────────────────────────────────────────────────────────

def detect_format(filepath):
    ext = Path(filepath).suffix.lower().lstrip(".")
    return ext if ext in SUPPORTED_IMPORT else None

def _normalize_row(row: dict, field_map: dict = None) -> dict:
    """Apply field mapping to normalize MES column names → OpsCore column names."""
    if not field_map:
        return row
    return {field_map.get(k, k): v for k, v in row.items()}

def import_csv(filepath, field_map=None):
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(_normalize_row(dict(row), field_map))
    return rows

def import_json(filepath, field_map=None):
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return [_normalize_row(r, field_map) for r in data]
    if isinstance(data, dict):
        # Handle common MES JSON envelope formats
        for key in ["records", "data", "shipments", "orders", "items", "rows", "results"]:
            if key in data:
                return [_normalize_row(r, field_map) for r in data[key]]
        return [_normalize_row(data, field_map)]
    return []

def import_xlsx(filepath, sheet=None, field_map=None):
    if not HAS_XLSX:
        raise ImportError("openpyxl required for Excel import. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.values)
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        result.append(_normalize_row(d, field_map))
    return result

def import_xml(filepath, record_tag=None, field_map=None):
    """
    Parse XML MES exports. Tries common record tag names automatically.
    Works with SAP iDocs, EDI XML, generic XML exports.
    """
    tree = ET.parse(filepath)
    root = tree.getroot()

    # Auto-detect record tag
    if not record_tag:
        common_tags = ["row", "record", "item", "shipment", "order",
                       "line", "entry", "IDOC", "E1EDL20", "E1EDL24"]
        for tag in common_tags:
            elements = root.findall(f".//{tag}")
            if elements:
                record_tag = tag
                break
        if not record_tag:
            # Use first-level children
            children = list(root)
            if children:
                record_tag = children[0].tag

    records = root.findall(f".//{record_tag}")
    result = []
    for rec in records:
        row = {}
        # Get attributes
        row.update(rec.attrib)
        # Get child elements as fields
        for child in rec:
            row[child.tag] = (child.text or "").strip()
        # Get own text
        if rec.text and rec.text.strip():
            row["value"] = rec.text.strip()
        result.append(_normalize_row(row, field_map))
    return result

def auto_import(filepath, field_map=None, sheet=None, record_tag=None):
    """Detect format and import automatically."""
    fmt = detect_format(filepath)
    if not fmt:
        raise ValueError(f"Unsupported file format: {Path(filepath).suffix}")

    dispatch = {
        "csv":  lambda: import_csv(filepath, field_map),
        "json": lambda: import_json(filepath, field_map),
        "xlsx": lambda: import_xlsx(filepath, sheet, field_map),
        "xml":  lambda: import_xml(filepath, record_tag, field_map),
    }
    records = dispatch[fmt]()
    print(f"[MES IMPORT] {fmt.upper()} → {len(records)} records from {Path(filepath).name}")
    return records

# ── EXPORT ─────────────────────────────────────────────────────────────────────

def export_csv(data: list, filepath: str):
    if not data:
        return
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    print(f"[MES EXPORT] CSV → {filepath} ({len(data)} records)")

def export_json(data, filepath: str, envelope_key="records"):
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    payload = {
        envelope_key: data,
        "exported_at": datetime.utcnow().isoformat(),
        "record_count": len(data) if isinstance(data, list) else 1,
        "source": "OpsCore"
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)
    print(f"[MES EXPORT] JSON → {filepath}")

def export_xlsx(data: list, filepath: str, sheet_name="OpsCore Export"):
    if not HAS_XLSX:
        raise ImportError("openpyxl required. Run: pip install openpyxl")
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        wb.save(filepath)
        return

    headers = list(data[0].keys())
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fgColor="1E293B", fill_type="solid")

    for row in data:
        ws.append([str(row.get(h, "")) for h in headers])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(filepath)
    print(f"[MES EXPORT] XLSX → {filepath} ({len(data)} rows)")

def export_xml(data: list, filepath: str, root_tag="OpsCore", record_tag="record"):
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    root = ET.Element(root_tag)
    root.set("exported_at", datetime.utcnow().isoformat())
    root.set("record_count", str(len(data)))
    root.set("source", "OpsCore")

    for row in data:
        rec = ET.SubElement(root, record_tag)
        for k, v in row.items():
            child = ET.SubElement(rec, str(k).replace(" ", "_"))
            child.text = str(v) if v is not None else ""

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    tree.write(filepath, encoding="unicode", xml_declaration=True)
    print(f"[MES EXPORT] XML → {filepath} ({len(data)} records)")

def auto_export(data, filepath: str, **kwargs):
    """Detect format from filepath extension and export."""
    fmt = Path(filepath).suffix.lower().lstrip(".")
    dispatch = {
        "csv":  lambda: export_csv(data, filepath),
        "json": lambda: export_json(data, filepath, **kwargs),
        "xlsx": lambda: export_xlsx(data, filepath, **kwargs),
        "xml":  lambda: export_xml(data, filepath, **kwargs),
    }
    if fmt not in dispatch:
        raise ValueError(f"Unsupported export format: .{fmt}")
    dispatch[fmt]()
    return filepath

# ── Field Map Presets (common MES → OpsCore mappings) ─────────────────────────

FIELD_MAPS = {
    # SAP SD/LE common field names → OpsCore shipment fields
    "sap_shipment": {
        "VBELN": "tracking_id",
        "KSCHL": "carrier",
        "LAND1": "destination",
        "WADAT": "ship_date",
        "LDDAT": "expected_date",
        "LFDAT": "actual_delivery_date",
        "GBSTA": "status",
    },
    # Generic WMS export
    "wms_generic": {
        "order_number":   "tracking_id",
        "ship_via":       "carrier",
        "ship_to_city":   "destination",
        "ship_date":      "ship_date",
        "promise_date":   "expected_date",
        "delivery_date":  "actual_delivery_date",
        "order_status":   "status",
    },
    # NetSuite
    "netsuite": {
        "Shipment Number":  "tracking_id",
        "Carrier":          "carrier",
        "Ship To":          "destination",
        "Ship Date":        "ship_date",
        "Estimated Delivery": "expected_date",
        "Actual Delivery":  "actual_delivery_date",
        "Status":           "status",
    },
}
